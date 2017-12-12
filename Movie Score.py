from openpyxl import Workbook
from openpyxl import load_workbook
import bs4
import sys
import operator
import time
import requests
from bs4 import BeautifulSoup
import urllib

def score(gross, budget, critic, audience, awards=0):

    total = ((gross-budget)*(((critic+audience)/2)/100))+awards
    return int(total)

##def dictionary(title, score):
##    output[title] = score
##    sorted_output = sorted(output.items(), key=operator.itemgetter(0), reverse=True)
##    for i in range(1):
##        print("i = ", i)
##        for j in range(1):
##            print("j = ", j)
##            ws.append([sorted_output[i][j]])
##            ws.append([sorted_output[i][j+1]])
##    wb.save(filename = "2016 Movie Scores.xlsx")
##    return sorted_output
##
##output = {}
##wb = Workbook()
##ws = wb.active


print("Arrival - "+str(score(203,47,94,82,120)))
print("Moonlight - "+str(score(65,4,98,80,250)))
print("Zootopia - "+str(score(1024,150,98,92,50)))
print("La La Land - "+str(score(446,30,92,81,380)))
print("Hell or High Water - "+str(score(38,12,98,88,40)))
print("Manchester by the Sea - "+str(score(75,9,95,77,140)))
print("Deadpool - "+str(score(783,58,84,90)))
print("Suicide Squad - "+str(score(746,175,25,61,50)))
print("Rogue One - "+str(score(1056,200,85,87,20)))
print("Captain America: Civil War - "+str(score(1153,250,90,89)))
print("Batman v Superman: Dawn of Justice - "+str(score(873,250,27,63)))
print("Doctor Strange - "+str(score(677,165,90,86,10)))
print("X-Men Apocalypse - "+str(score(543,178,48,66)))
print("Ghostbusters - "+str(score(229,144,73,52)))
print("Chairman of the Board - "+str(score(0.18,10,13,18)))
print("Waterworld - "+str(score(264,175,42,43)))
print("Avatar - "+str(score(2788,237,84,82)))
print("Titanic - "+str(score(2186,200,88,69)))
print("The Force Awakens - "+str(score(2068,245,93,89)))
##
##print(dictionary("Arrival", score(203,47,94,82,120)))
##print(dictionary("Moonlight", score(65,4,98,80,250)))
##print(load_workbook("2016 Movie Scores.xlsx"))

def critic(rt_url):
    response = requests.get(rt_url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    content_div = soup.find(id="all-critics-numbers").find(class_="meter-value superPageFontColor")
    
    for element in content_div.find_all("span", recursive=False):
        
        critic = element.get_text()
                
        return int(critic)

def audience(rt_url):
    response = requests.get(rt_url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    content_div = soup.find(id="topSection").find(class_="media-body")
    sub_div = content_div.find(class_="meter-value")
    
    for element in sub_div.find_all("span", recursive=False):
        
        audience = element.get_text()[:-1]        

    return int(audience)

def gross(bo_url):
    response = requests.get(bo_url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    content_div = soup.find(class_="mp_box_content").table
    box_office = content_div.get_text()
    ww_gross = (box_office[box_office.find("Worldwide:")+13:])
    ww_gross = ww_gross.replace(",","")
    return round((float(ww_gross)/1000000),0)

def budget(bo_url):
    response = requests.get(bo_url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    content = soup.get_text()
    budget = content[content.find("Production Budget:"):content.find(" million")]
    budget = budget.replace("Production Budget: $", "")
    return float(budget)

def awards(imdb_url):
    response = requests.get(imdb_url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    content_div = soup.find(id="main").find(class_="article listo").find("table")
    content = content_div.get_text()
    win_text = content[:content.find("Nominated")]
    nom_text = content[content.find("Nominated"):]
    wins = win_text.count("Best")
    bp_bonus = 0
    if win_text.find("Best Picture") == True:
        bp_bonus += 50
    noms = nom_text.count("Best")
    return (wins*50)+(noms*10)+bp_bonus

rt_url = "https://www.rottentomatoes.com/m/1084146_thin_red_line"
bo_url = "http://www.boxofficemojo.com/movies/?id=thinredline.htm"
imdb_url = "http://www.imdb.com/title/tt0120863/awards?ref_=tt_awd"



##print("Movie: The Thin Red Line")
##print("Box Office: $", int(gross(bo_url)), " million")
##print("Budget: $", int(budget(bo_url)), " million")
##print("Critic Score: ", critic(rt_url), "%")
##print("Audience Score: ", audience(rt_url), "%")
##print("Oscars Bonus: ", awards(imdb_url))
##print("Total Score: ", score(gross(bo_url), budget(bo_url), critic(rt_url), audience(rt_url), awards(imdb_url)))
